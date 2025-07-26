from django.forms import ValidationError
from django.shortcuts import render

import os
# import tempfile
from django.http import FileResponse, HttpResponse, HttpResponseNotFound
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
from datetime import datetime
# from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
import json
from rest_framework.decorators import api_view
from openpyxl.styles import Font
from .models import InvoiceDP, InvoiceFinal, KwitansiDP, KwitansiFinal, TemplateProposal, ProposalTemplateHistory, KontrakTemplateHistory, BAST
from django.contrib.auth import get_user_model
from django.http import JsonResponse
# from django.contrib.auth.decorators import login_required
# from pptx import Presentation
from django.core.files.storage import default_storage
from decimal import Decimal
from .serializers import TemplateProposalSerializer, ProposalTemplateHistorySerializer
import aspose.slides as slides
import aspose.pydrawing as drawing
import io


User = get_user_model()

def parse_date(date_str):
    if not date_str:
        raise ValidationError("Tanggal tidak boleh kosong")
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        raise ValidationError(f"Format tanggal salah: {date_str}")

def get_next_bast_number():
    from .models import BAST
    latest = BAST.objects.order_by('-created_at').first()
    if latest and latest.nomor:
        try:
            return int(latest.nomor.split("/")[0]) + 1
        except:
            return 1
    return 1

@api_view(['POST'])
def export_existing_bast(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        judul_survei = data.get('data')  # Ambil dari payload: {data: "judul survei"}

        try:
            bast_obj = BAST.objects.get(judul_survei=judul_survei)
        except BAST.DoesNotExist:
            return JsonResponse({'error': 'Data BAST tidak ditemukan untuk judul survei tersebut.'}, status=404)
    
    # Ambil data dari objek model
    bast_data = {
        'nomor': bast_obj.nomor,
        'nama_pihak_pertama': bast_obj.nama_pihak_pertama,
        'jabatan_pihak_pertama': bast_obj.jabatan_pihak_pertama,
        'alamat_pihak_pertama': bast_obj.alamat_pihak_pertama,
        'nama_pihak_kedua': bast_obj.nama_pihak_kedua,
        'jabatan_pihak_kedua': bast_obj.jabatan_pihak_kedua,
        'alamat_pihak_kedua': bast_obj.alamat_pihak_kedua,
        'nomor_spk': bast_obj.nomor_spk,
        'tanggal_spk': bast_obj.tanggal_spk.strftime('%Y-%m-%d') if bast_obj.tanggal_spk else '',
        'judul_survei': bast_obj.judul_survei,
        'nomor_addendum': bast_obj.nomor_addendum,
        'tanggal_addendum': bast_obj.tanggal_addendum.strftime('%Y-%m-%d') if bast_obj.tanggal_addendum else '',
        'tanggal_tertulis': bast_obj.tanggal_tertulis,
        'tanggal_serah_terima': bast_obj.tanggal.strftime('%Y-%m-%d') if bast_obj.tanggal else datetime.now().strftime('%Y-%m-%d'),
        'nilai_kontrak_angka': str(bast_obj.nilai_kontrak_angka),
        'nilai_kontrak_tertulis': bast_obj.nilai_kontrak_tertulis,
    }


    # current_date = datetime.now()
    # month_roman = month_to_roman(current_date.month)
    # year = current_date.year
    # bast_number = get_next_bast_number()
    # bast_code = f"{bast_number}/BAST/{month_roman}/{year}"
    # bast_id = bast_code

    # formatted_date = datetime.strptime(
    #     bast_data['tanggal_serah_terima'], '%Y-%m-%d'
    # ).strftime("Jakarta, %d %B %Y")

    baris_atas, baris_bawah = split_to_two_cells(bast_data['judul_survei'], max_length=60)

    formatted_nilai_kontrak = f"{bast_data['nilai_kontrak_angka']} (sudah termasuk pajak)"
    formatted_tanggal_tertulis = f"{bast_data['tanggal_tertulis']} kami yang bertandatangan di bawah ini:"


# Load template Excel
    template_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/templates/templateBAST.xlsx')
    workbook = load_workbook(template_path)
    sheet = workbook.active

    # Styling font
    times_font = Font(name="Times New Roman")
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = times_font


    # Isi data ke template
    sheet['B1'].font = Font(name="Times New Roman", bold=True, underline="single")
    sheet['B25'].font = Font(name="Times New Roman", bold=True)
    sheet['C25'].font = Font(name="Times New Roman", bold=True)
    sheet['B2'] = bast_data['nomor']
    sheet['D3'] = formatted_tanggal_tertulis
    sheet['G5'] = bast_data['nama_pihak_pertama']
    sheet['G6'] = bast_data['jabatan_pihak_pertama']
    sheet['G7'] = bast_data['alamat_pihak_pertama']
    sheet['G9'] = bast_data['nama_pihak_kedua']
    sheet['G10'] = bast_data['jabatan_pihak_kedua']
    sheet['G11'] = bast_data['alamat_pihak_kedua']
    sheet['I14'] = bast_data['nomor_spk']
    sheet['I15'] = bast_data['tanggal_spk']
    sheet['I17'] = bast_data['nomor_addendum']
    sheet['I18'] = bast_data['tanggal_addendum']
    sheet['I20'] = bast_data['nomor']
    sheet['I21'] = bast_data['tanggal_serah_terima']
    sheet['J23'] = baris_atas
    sheet['B24'] = baris_bawah
    sheet['C26'] = bast_data['judul_survei']
    sheet['E28'] = formatted_nilai_kontrak
    sheet['D29'] = bast_data['nilai_kontrak_tertulis']
    sheet['D37'] = bast_data['nama_pihak_kedua']
    sheet['D38'] = bast_data['jabatan_pihak_kedua']
    sheet['J37'] = bast_data['nama_pihak_pertama']
    sheet['J38'] = bast_data['jabatan_pihak_pertama']
    # sheet['G10'] = formatted_date  # misalnya tanda tangan Jakarta, tgl

    # Output response
    filename = f"BAST_{bast_data['nomor']}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    workbook.save(response)
    return response


@api_view(['POST'])
def generate_bast(request):
    if request.method == 'POST':
        data = json.loads(request.body)

        # Persiapkan data
        bast_data = {
            'nama_pihak_pertama': data.get('nama_pihak_pertama', ''),
            'jabatan_pihak_pertama': data.get('jabatan_pihak_pertama', ''),
            'alamat_pihak_pertama': data.get('alamat_pihak_pertama', ''),
            'nama_pihak_kedua': data.get('nama_pihak_kedua', ''),
            'jabatan_pihak_kedua': data.get('jabatan_pihak_kedua', ''),
            'alamat_pihak_kedua': data.get('alamat_pihak_kedua', ''),
            'nomor_spk': data.get('nomor_spk', ''),
            'tanggal_spk': data.get('tanggal_spk', ''),
            'judul_survei': data.get('judul_survei', ''),
            'nomor_addendum': data.get('nomor_addendum', ''),
            'tanggal_addendum': data.get('tanggal_addendum', ''),
            'tanggal_tertulis': data.get('tanggal_tertulis', ''),
            'tanggal_serah_terima': data.get('tanggal', datetime.now().strftime('%Y-%m-%d')),
            'nilai_kontrak_angka': data.get('nilai_kontrak_angka', '0'),
            'nilai_kontrak_tertulis': data.get('nilai_kontrak_tertulis', ''),
        }

        print(f"BAST Data: {bast_data}")

        current_date = datetime.now()
        month_roman = month_to_roman(current_date.month)
        year = current_date.year
        bast_number = get_next_bast_number()
        bast_code = f"{bast_number}/BAST/{month_roman}/{year}"
        bast_id = bast_code

        formatted_date = datetime.strptime(
            bast_data['tanggal_serah_terima'], '%Y-%m-%d'
        ).strftime("Jakarta, %d %B %Y")

        baris_atas, baris_bawah = split_to_two_cells(bast_data['judul_survei'], max_length=60)

        formatted_nilai_kontrak = f"{bast_data['nilai_kontrak_angka']} (sudah termasuk pajak)"
        formatted_tanggal_tertulis = f"{bast_data['tanggal_tertulis']} kami yang bertandatangan di bawah ini:"

        # Simpan data jika belum ada
        if not BAST.objects.filter(nomor=bast_id).exists():
            BAST.objects.create(
                nomor=bast_id,
                tanggal=bast_data['tanggal_serah_terima'],
                nama_pihak_pertama=bast_data['nama_pihak_pertama'],
                jabatan_pihak_pertama=bast_data['jabatan_pihak_pertama'],
                alamat_pihak_pertama=bast_data['alamat_pihak_pertama'],
                nama_pihak_kedua=bast_data['nama_pihak_kedua'],
                jabatan_pihak_kedua=bast_data['jabatan_pihak_kedua'],
                alamat_pihak_kedua=bast_data['alamat_pihak_kedua'],
                nomor_spk=bast_data['nomor_spk'],
                tanggal_spk=bast_data['tanggal_spk'],
                judul_survei=bast_data['judul_survei'],
                nomor_addendum=bast_data['nomor_addendum'],
                tanggal_addendum=bast_data['tanggal_addendum'],
                tanggal_tertulis=bast_data['tanggal_tertulis'],
                nilai_kontrak_angka=bast_data['nilai_kontrak_angka'],
                nilai_kontrak_tertulis=bast_data['nilai_kontrak_tertulis'],
            )
        else:
            print(f"Record with id {bast_id} already exists. Skipping creation.")

        # Load template Excel
        template_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/templates/templateBAST.xlsx')
        workbook = load_workbook(template_path)
        sheet = workbook.active

        # Styling font
        times_font = Font(name="Times New Roman")
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = times_font


        # Isi data ke template
        sheet['B1'].font = Font(name="Times New Roman", bold=True, underline="single")
        sheet['B25'].font = Font(name="Times New Roman", bold=True)
        sheet['C25'].font = Font(name="Times New Roman", bold=True)
        sheet['B2'] = bast_code
        sheet['D3'] = formatted_tanggal_tertulis
        sheet['G5'] = bast_data['nama_pihak_pertama']
        sheet['G6'] = bast_data['jabatan_pihak_pertama']
        sheet['G7'] = bast_data['alamat_pihak_pertama']
        sheet['G9'] = bast_data['nama_pihak_kedua']
        sheet['G10'] = bast_data['jabatan_pihak_kedua']
        sheet['G11'] = bast_data['alamat_pihak_kedua']
        sheet['I14'] = bast_data['nomor_spk']
        sheet['I15'] = bast_data['tanggal_spk']
        sheet['I17'] = bast_data['nomor_addendum']
        sheet['I18'] = bast_data['tanggal_addendum']
        sheet['I20'] = bast_code
        sheet['I21'] = bast_data['tanggal_serah_terima']
        sheet['J23'] = baris_atas
        sheet['B24'] = baris_bawah
        sheet['C26'] = bast_data['judul_survei']
        sheet['E28'] = formatted_nilai_kontrak
        sheet['D29'] = bast_data['nilai_kontrak_tertulis']
        sheet['D37'] = bast_data['nama_pihak_kedua']
        sheet['D38'] = bast_data['jabatan_pihak_kedua']
        sheet['J37'] = bast_data['nama_pihak_pertama']
        sheet['J38'] = bast_data['jabatan_pihak_pertama']
        # sheet['G10'] = formatted_date  # misalnya tanda tangan Jakarta, tgl

        # Output response
        filename = f"BAST_{bast_code}.xlsx"
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        workbook.save(response)
        return response

def split_to_two_cells(text, max_length=60):
    if len(text) <= max_length:
        return text, ""
    else:
        # Cari spasi terdekat agar tidak memotong di tengah kata
        split_index = text.rfind(" ", 0, max_length)
        if split_index == -1:
            split_index = max_length
        return text[:split_index], text[split_index+1:]

@api_view(['GET'])
def convert_pptx_to_image(request):
    folder_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/templates/proposal/')

    # Cari semua file .pptx dalam folder
    pptx_files = [os.path.join(folder_path, f) for f in os.listdir(folder_path) if f.endswith('.pptx')]

    if not pptx_files:
        return HttpResponseNotFound("No proposal templates found.")

    # Ambil file terbaru berdasarkan waktu modifikasi
    latest_file = max(pptx_files, key=os.path.getmtime)

    # Gunakan Aspose.Slides untuk membuka dan convert slide pertama
    with slides.Presentation(latest_file) as presentation:
        slide = presentation.slides[0]
        bmp = slide.get_thumbnail(2.0, 2.0)  # Scale 2x

        image_stream = io.BytesIO()
        bmp.save(image_stream, drawing.imaging.ImageFormat.png)
        image_stream.seek(0)

        return FileResponse(image_stream, content_type='image/png')

@api_view(['POST'])
def upload_template_proposal(request):
    file = request.FILES.get('template')
    if not file:
        return JsonResponse({'error': 'File not provided'}, status=400)

    # Buat record history terlebih dahulu (tanpa file)
    history = ProposalTemplateHistory.objects.create(uploaded_by=request.user)

    # Simpan file ke folder manual
    history_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/templates/proposal')
    os.makedirs(history_path, exist_ok=True)

    # Tentukan path file manual
    file_path = os.path.join(history_path, f'templateProposal_v{history.id}.pptx')

    with open(file_path, 'wb+') as destination:
        for chunk in file.chunks():
            destination.write(chunk)

    return JsonResponse({'message': 'Template uploaded', 'id': history.id}, status=201)

@api_view(['DELETE'])
def delete_template_proposal_by_id(request, id):
    try:
        template = ProposalTemplateHistory.objects.get(id=id)
        template.file.delete(save=False)
        template.delete()
        return JsonResponse({"message": "Template deleted successfully."}, status=200)
    except ProposalTemplateHistory.DoesNotExist:
        return JsonResponse({"error": "Template not found."}, status=404)

@api_view(['GET'])
def get_proposal_template_history(request):
    templates = ProposalTemplateHistory.objects.all().order_by('-uploaded_at')
    serializer = ProposalTemplateHistorySerializer(templates, many=True)
    return JsonResponse(serializer.data, safe=False)

@api_view(['GET'])
def list_template_proposals(request):
    templates = TemplateProposal.objects.order_by('-uploaded_at')
    serializer = TemplateProposalSerializer(templates, many=True)
    return JsonResponse(serializer.data)

@api_view(['GET'])
def get_kontrak_template_history(request):
    templates = KontrakTemplateHistory.objects.all().order_by('-uploaded_at')
    serializer = ProposalTemplateHistorySerializer(templates, many=True)  # Reuse serializer if identical
    return JsonResponse(serializer.data, safe=False)

@api_view(['DELETE'])
def delete_kontrak_template(request, id):
    try:
        template = KontrakTemplateHistory.objects.get(id=id)
        template.file.delete(save=False)
        template.delete()
        return JsonResponse({"message": "Template deleted successfully."}, status=200)
    except KontrakTemplateHistory.DoesNotExist:
        return JsonResponse({"error": "Template not found."}, status=404)


@api_view(['GET'])
def download_template_kontrak(request):
    template_id = request.GET.get('id')
    if not template_id:
        return JsonResponse({'error': 'Template ID is required'}, status=400)

    filepath = os.path.join(settings.BASE_DIR, 'dokumen_pendukung', 'templates', 'kontrak', f'templateKontrak_v{template_id}.docx')
    if os.path.exists(filepath):
        return FileResponse(open(filepath, 'rb'), content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
    return HttpResponseNotFound("Template not found")

@api_view(['POST'])
def upload_template_kontrak(request):
    file = request.FILES.get('template')
    if not file or not file.name.endswith('.docx'):
        return JsonResponse({'error': 'Only .docx files are allowed'}, status=400)

    template = KontrakTemplateHistory.objects.create(file=file, uploaded_by=request.user)

    # Simpan file di folder `dokumen_pendukung/templates/kontrak/templateKontrak_v{id}.docx`
    kontrak_folder = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/templates/kontrak/')
    os.makedirs(kontrak_folder, exist_ok=True)

    versioned_path = os.path.join(kontrak_folder, f'templateKontrak_v{template.id}.docx')
    with open(versioned_path, 'wb+') as dest:
        for chunk in file.chunks():
            dest.write(chunk)

    return JsonResponse({'message': 'Template kontrak uploaded', 'id': template.id}, status=201)

@api_view(['GET']) 
def download_template_proposal(request):
    template_id = request.GET.get('id')
    if not template_id:
        return JsonResponse({'error': 'Template ID is required'}, status=400)

    # Tentukan path file berdasarkan id
    filename = f'templateProposal_v{template_id}.pptx'
    file_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung', 'templates', 'proposal', filename)

    if os.path.exists(file_path):
        response = FileResponse(open(file_path, 'rb'), content_type='application/vnd.openxmlformats-officedocument.presentationml.presentation')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    else:
        return HttpResponseNotFound("Template file not found.")


# Helper function to convert month to Roman numeral
def month_to_roman(month):
    roman_numerals = {
        1: "I", 2: "II", 3: "III", 4: "IV",
        5: "V", 6: "VI", 7: "VII", 8: "VIII",
        9: "IX", 10: "X", 11: "XI", 12: "XII"
    }
    return roman_numerals.get(month, "")

# Helper function to get the next invoice number
def get_next_invoice_number():
    current_date = datetime.now()
    year = current_date.year
    counter_file_path = os.path.join(settings.BASE_DIR, 'invoice_{year}_counter.txt')

    # Check if the file exists, and create it if not
    if not os.path.exists(counter_file_path):
        with open(counter_file_path, 'w') as file:
            file.write("0")  # Initialize counter at 0

    with open(counter_file_path, 'r') as file:
        file_number = int(file.read())

    # Ambil nomor invoice terakhir dari InvoiceDP
    latest_dp = (
        InvoiceDP.objects
        .filter(id__endswith=f'/{year}')
        .order_by('-id')
        .first()
    )
    dp_number = int(latest_dp.id.split('/')[0]) if latest_dp else 0

    # Ambil nomor invoice terakhir dari InvoiceFinal
    latest_final = (
        InvoiceFinal.objects
        .filter(id__endswith=f'/{year}')
        .order_by('-id')
        .first()
    )
    final_number = int(latest_final.id.split('/')[0]) if latest_final else 0

    # Ambil yang terbesar dari dua tabel
    db_number = max(dp_number, final_number)

    # Logika penyesuaian counter
    if file_number == db_number:
        # Sinkron, generate baru
        new_number = file_number + 1
        with open(counter_file_path, 'w') as file:
            file.write(str(new_number))
        return f"{new_number:03}"
    elif file_number > db_number:
        # File lebih tinggi dari DB, gunakan saja
        return f"{file_number:03}"
    else:
        # File lebih rendah dari DB, update file agar sinkron
        new_number = db_number + 1
        with open(counter_file_path, 'w') as file:
            file.write(str(new_number))
        return f"{new_number:03}"


# Helper function to get the next kwitansi number
def get_next_kwitansi_number():
    current_date = datetime.now()
    year = current_date.year
    counter_file_path = os.path.join(settings.BASE_DIR, 'kwitansi_{year}_counter.txt')

    # Check if the file exists, and create it if not
    if not os.path.exists(counter_file_path):
        with open(counter_file_path, 'w') as file:
            file.write("0")  # Initialize counter at 0

    # Read the current counter value
    with open(counter_file_path, 'r') as file:
        last_number = int(file.read())

    # Ambil nomor terakhir dari KwitansiDP
    latest_dp = (
        KwitansiDP.objects
        .filter(id__endswith=f'/{year}')
        .order_by('-id')
        .first()
    )
    dp_number = int(latest_dp.id.split('/')[0]) if latest_dp else 0

    # Ambil nomor terakhir dari KwitansiFinal
    latest_final = (
        KwitansiFinal.objects
        .filter(id__endswith=f'/{year}')
        .order_by('-id')
        .first()
    )
    final_number = int(latest_final.id.split('/')[0]) if latest_final else 0

    # Ambil nomor terbesar dari database
    db_number = max(dp_number, final_number)

    # Sinkronisasi counter
    if last_number == db_number:
        # Sinkron, generate baru
        new_number = last_number + 1
        with open(counter_file_path, 'w') as file:
            file.write(str(new_number))
        return f"{new_number:03}"
    elif last_number > db_number:
        # File lebih tinggi, pakai itu
        return f"{last_number:03}"
    else:
        # File lebih rendah dari DB, sinkronkan
        new_number = db_number + 1
        with open(counter_file_path, 'w') as file:
            file.write(str(new_number))
        return f"{new_number:03}"



@api_view(['POST'])
def generate_invoice_dp(request):
    if request.method == 'POST':
        # Parse JSON data from the request body
        data = json.loads(request.body)
    
    user_data = {
        'client_name': data.get('client_name', 'Default Client'),
        'survey_name': data.get('survey_name', 'Default Survey'),
        'respondent_count': data.get('respondent_count', '0'),
        'address': data.get('address', 'Default Address'),
        'amount': data.get('amount', '0'),
        'paid_percentage': data.get('paid_percentage', '60'),
        'nominal_tertulis' : data.get('nominal_tertulis', ''),
        'additional_info': data.get('additional_info', 'No additional info'),
        'date': data.get('date', datetime.now().strftime('%Y-%m-%d')),
        'beneficiary_bank_name': data.get('beneficiary_bank_name', ''),
        'beneficiary_account_name': data.get('beneficiary_account_name', ''),
        'beneficiary_account_number': data.get('beneficiary_account_number', ''),
        'account_currency': data.get('account_currency', ''),
        'beneficiary_bank_address': data.get('beneficiary_bank_address', ''),
        'beneficiary_swift_code': data.get('beneficiary_swift_code', ''),
        'tax_id': data.get('tax_id', ''),
    }


    # Generate unique invoice code
    current_date = datetime.now()
    month_roman = month_to_roman(current_date.month)
    year = current_date.year
    invoice_number = get_next_invoice_number()
    invoice_code = f"Inv No: {invoice_number}/SURV/LSI/{month_roman}/{year}"
    invoice_id = f"{invoice_number}/SURV/LSI/{month_roman}/{year}"
    doc = "invoiceDP"

    if Decimal(user_data['amount']) > Decimal('999999999999.99'):
        return HttpResponse({"error": "Amount terlalu besar untuk disimpan."}, status=400)

    # Save the data to the invoice_dp table
    elif not InvoiceDP.objects.filter(id=invoice_id).exists(): 
        InvoiceDP.objects.create(
            id=invoice_id,
            client_name=user_data['client_name'],
            survey_name=user_data['survey_name'],
            respondent_count=user_data['respondent_count'],
            address=user_data['address'],
            amount=user_data['amount'],
            nominal_tertulis=user_data['nominal_tertulis'],
            paid_percentage=user_data['paid_percentage'],
            additional_info=user_data['additional_info'],
            date=user_data['date'],
            doc_type=doc,
            beneficiary_bank_name=user_data['beneficiary_bank_name'],
            beneficiary_account_name=user_data['beneficiary_account_name'],
            beneficiary_account_number=user_data['beneficiary_account_number'],
            account_currency=user_data['account_currency'],
            beneficiary_bank_address=user_data['beneficiary_bank_address'],
            beneficiary_swift_code=user_data['beneficiary_swift_code'],
            tax_id=user_data['tax_id'],
        )
    else:
        print(f"Record with id {invoice_id} already exists. Skipping creation.")

    # Load the Excel template
    template_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/templates/templateInvoiceDP.xlsx')

    workbook = load_workbook(template_path)
    sheet = workbook.active
    respondent_message = f"sampel {user_data['respondent_count']} responden"
    paid_percentage_message = f"{user_data['paid_percentage']}%"
    formatted_date = datetime.strptime(user_data['date'], "%Y-%m-%d").strftime("%d %B %Y")
    date_message = f"Date: {formatted_date}"
    sheet.merge_cells('B31:E31')
    sheet.merge_cells('C16:F16')
    sheet.merge_cells('B27:E27')
    sheet.merge_cells('B28:E28')
    sheet.merge_cells('G14:H14')

    # Fill out the required cells in the Excel file with user data
    sheet['A9'] = invoice_code
    sheet['C16'] = user_data['client_name']           
    sheet['B27'] = user_data['survey_name']           
    sheet['B31'] = respondent_message      
    sheet['C18'] = user_data['address']               
    sheet['G27'] = user_data['amount']                
    sheet['C35'] = user_data['nominal_tertulis']
    sheet['C35'].font = Font(name="Times New Roman", bold=True, underline="single")
    sheet['F27'] = paid_percentage_message      
    sheet['B28'] = user_data['additional_info']       
    sheet['G14'] = date_message    
    sheet['C42'] = f": {user_data['beneficiary_bank_name']}"
    sheet['C39'] = f": {user_data['beneficiary_account_name']}"
    sheet['C40'] = f": {user_data['beneficiary_account_number']}"
    sheet['C40'].font = Font(name="Times New Roman", bold=True)
    sheet['C41'] = f": {user_data['account_currency']}"
    sheet['C43'] = f": {user_data['beneficiary_bank_address']}"
    sheet['C44'] = f": {user_data['beneficiary_swift_code']}"
    sheet['C45'] = f": {user_data['tax_id']}"          

    # Path to the image you want to add
    header_image_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/images/header.png')  
    header_img = Image(header_image_path)
    header_img.width, header_img.height = 846.6, 136  

    # Path to the image you want to add
    invoice_image_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/images/invoice.png') 
    invoice_img = Image(invoice_image_path)
    invoice_img.width, invoice_img.height = 275.9, 75.59

    # # Path to the image you want to add
    # ttd_image_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/images/ttd.png')
    # ttd_img = Image(ttd_image_path)
    # ttd_img.width, ttd_img.height = 314.83, 173.48

    # Add image to the specified cell location
    sheet.add_image(header_img, 'A1') 
    sheet.add_image(invoice_img, 'G8')

    # Generate a filename
    filename = f"{user_data['survey_name']}_invoiceDP_{invoice_code}.xlsx"

    # Prepare the response as an Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'

    # Save workbook to the response
    workbook.save(response)
    return response

    return HttpResponse(status=405)  # Method not allowed for non-POST requests

@api_view(['POST'])
def export_existing_invoice_dp(request):
    if request.method == 'POST':
        # Parse JSON data from the request body
        data = json.loads(request.body)
    
    user_data = {
        'invoice_code': data.get('id', ''),
        'client_name': data.get('client_name', 'Default Client'),
        'survey_name': data.get('survey_name', 'Default Survey'),
        'respondent_count': data.get('respondent_count', '0'),
        'address': data.get('address', 'Default Address'),
        'amount': data.get('amount', '0'),
        'paid_percentage': data.get('paid_percentage', '60'),
        'nominal_tertulis' : data.get('nominal_tertulis', ''),
        'additional_info': data.get('additional_info', 'No additional info'),
        'date': data.get('date', datetime.now().strftime('%Y-%m-%d')),
    }


    # Load the Excel template
    template_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/templates/templateInvoiceDP.xlsx')

    workbook = load_workbook(template_path)
    sheet = workbook.active
    respondent_message = f"sampel {user_data['respondent_count']} responden"
    paid_percentage_message = f"{user_data['paid_percentage']}%"
    formatted_date = datetime.strptime(user_data['date'], "%Y-%m-%d").strftime("%d %B %Y")
    date_message = f"Date: {formatted_date}"
    sheet.merge_cells('B31:E31')
    sheet.merge_cells('C16:F16')
    sheet.merge_cells('B27:E27')
    sheet.merge_cells('B28:E28')
    sheet.merge_cells('G14:H14')

    # Fill out the required cells in the Excel file with user data
    sheet['A9'] = user_data['invoice_code']
    sheet['C16'] = user_data['client_name']           
    sheet['B27'] = user_data['survey_name']           
    sheet['B31'] = respondent_message      
    sheet['C18'] = user_data['address']               
    sheet['G27'] = user_data['amount']                
    sheet['C35'] = user_data['nominal_tertulis']
    sheet['C35'].font = Font(name="Times New Roman", bold=True, underline="single")
    sheet['F27'] = paid_percentage_message      
    sheet['B28'] = user_data['additional_info']       
    sheet['G14'] = date_message                 

    # Path to the image you want to add
    header_image_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/images/header.png')  
    header_img = Image(header_image_path)
    header_img.width, header_img.height = 846.6, 136  

    # Path to the image you want to add
    invoice_image_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/images/invoice.png') 
    invoice_img = Image(invoice_image_path)
    invoice_img.width, invoice_img.height = 275.9, 75.59

    # Path to the image you want to add
    ttd_image_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/images/ttd.png')
    ttd_img = Image(ttd_image_path)
    ttd_img.width, ttd_img.height = 314.83, 173.48

    # Add image to the specified cell location
    sheet.add_image(header_img, 'A1') 
    sheet.add_image(invoice_img, 'G8')
    # sheet.add_image(ttd_img, 'G37')

    invoice_id = f"Inv No: {user_data['invoice_code']}"

    # Generate a filename
    filename = f"{user_data['survey_name']}_invoiceDP_{invoice_id}.xlsx"

    # Prepare the response as an Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'

    # Save workbook to the response
    workbook.save(response)
    return response

    return HttpResponse(status=405)  # Method not allowed for non-POST requests

@api_view(['POST'])
def export_existing_invoice_final(request):
    if request.method == 'POST':
        # Parse JSON data from the request body
        data = json.loads(request.body)
    
    user_data = {
        'invoice_code': data.get('id', ''),
        'client_name': data.get('client_name', 'Default Client'),
        'survey_name': data.get('survey_name', 'Default Survey'),
        'respondent_count': data.get('respondent_count', '0'),
        'address': data.get('address', 'Default Address'),
        'amount': data.get('amount', '0'),
        'paid_percentage': data.get('paid_percentage', '60'),
        'nominal_tertulis' : data.get('nominal_tertulis', ''),
        'additional_info': data.get('additional_info', 'No additional info'),
        'date': data.get('date', datetime.now().strftime('%Y-%m-%d')),
    }


    # Load the Excel template
    template_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/templates/templateInvoiceFinal.xlsx')

    workbook = load_workbook(template_path)
    sheet = workbook.active
    respondent_message = f"sampel {user_data['respondent_count']} responden"
    paid_percentage_message = f"{user_data['paid_percentage']}%"
    formatted_date = datetime.strptime(user_data['date'], "%Y-%m-%d").strftime("%d %B %Y")
    date_message = f"Date: {formatted_date}"
    sheet.merge_cells('B31:E31')
    sheet.merge_cells('C16:F16')
    sheet.merge_cells('B27:E27')
    sheet.merge_cells('B28:E28')
    sheet.merge_cells('G14:H14')

    # Fill out the required cells in the Excel file with user data
    sheet['A9'] = user_data['invoice_code']
    sheet['C16'] = user_data['client_name']           
    sheet['B27'] = user_data['survey_name']           
    sheet['B31'] = respondent_message      
    sheet['C18'] = user_data['address']               
    sheet['G27'] = user_data['amount']                
    sheet['C35'] = user_data['nominal_tertulis']
    sheet['C35'].font = Font(name="Times New Roman", bold=True, underline="single")
    sheet['F27'] = paid_percentage_message      
    sheet['B28'] = user_data['additional_info']       
    sheet['G14'] = date_message                 

    # Path to the image you want to add
    header_image_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/images/header.png')  
    header_img = Image(header_image_path)
    header_img.width, header_img.height = 846.6, 136  

    # Path to the image you want to add
    invoice_image_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/images/invoice.png') 
    invoice_img = Image(invoice_image_path)
    invoice_img.width, invoice_img.height = 275.9, 75.59

    # Path to the image you want to add
    ttd_image_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/images/ttd.png')
    ttd_img = Image(ttd_image_path)
    ttd_img.width, ttd_img.height = 314.83, 173.48

    # Add image to the specified cell location
    sheet.add_image(header_img, 'A1') 
    sheet.add_image(invoice_img, 'G8')
    sheet.add_image(ttd_img, 'G37')

    invoice_id = f"Inv No: {user_data['invoice_code']}"

    # Generate a filename
    filename = f"{user_data['survey_name']}_invoiceFinal_{invoice_id}.xlsx"

    # Prepare the response as an Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'

    # Save workbook to the response
    workbook.save(response)
    return response

    return HttpResponse(status=405)  # Method not allowed for non-POST requests

@api_view(['POST'])
def generate_invoice_final(request):
    if request.method == 'POST':
        # Parse JSON data from the request body
        data = json.loads(request.body)
   
    user_data = {
        'client_name': data.get('client_name', 'Default Client'),
        'survey_name': data.get('survey_name', 'Default Survey'),
        'respondent_count': data.get('respondent_count', '0'),
        'address': data.get('address', 'Default Address'),
        'amount': data.get('amount', '0'),
        'paid_percentage': data.get('paid_percentage', '60'),
        'nominal_tertulis' : data.get('nominal_tertulis', ''),
        'additional_info': data.get('additional_info', 'No additional info'),
        'date': data.get('date', datetime.now().strftime('%Y-%m-%d')),
        'beneficiary_bank_name': data.get('beneficiary_bank_name', ''),
        'beneficiary_account_name': data.get('beneficiary_account_name', ''),
        'beneficiary_account_number': data.get('beneficiary_account_number', ''),
        'account_currency': data.get('account_currency', ''),
        'beneficiary_bank_address': data.get('beneficiary_bank_address', ''),
        'beneficiary_swift_code': data.get('beneficiary_swift_code', ''),
        'tax_id': data.get('tax_id', ''),
    }


    # Generate unique invoice code
    current_date = datetime.now()
    month_roman = month_to_roman(current_date.month)
    year = current_date.year
    invoice_number = get_next_invoice_number()
    invoice_code = f"Inv No: {invoice_number}/SURV/LSI/{month_roman}/{year}"
    invoice_id = f"{invoice_number}/SURV/LSI/{month_roman}/{year}"
    doc = "invoiceFinal"

    if Decimal(user_data['amount']) > Decimal('999999999999.99'):
        return HttpResponse({"error": "Amount terlalu besar untuk disimpan."}, status=400)

    # Save the data to the invoice_final table
    elif not InvoiceFinal.objects.filter(id=invoice_id).exists(): 
        InvoiceFinal.objects.create(
            id=invoice_id,
            client_name=user_data['client_name'],
            survey_name=user_data['survey_name'],
            respondent_count=user_data['respondent_count'],
            address=user_data['address'],
            amount=user_data['amount'],
            nominal_tertulis=user_data['nominal_tertulis'],
            paid_percentage=user_data['paid_percentage'],
            additional_info=user_data['additional_info'],
            date=user_data['date'],
            doc_type=doc,
            beneficiary_bank_name=user_data['beneficiary_bank_name'],
            beneficiary_account_name=user_data['beneficiary_account_name'],
            beneficiary_account_number=user_data['beneficiary_account_number'],
            account_currency=user_data['account_currency'],
            beneficiary_bank_address=user_data['beneficiary_bank_address'],
            beneficiary_swift_code=user_data['beneficiary_swift_code'],
            tax_id=user_data['tax_id'],
        )
    else:
        print(f"Record with id {invoice_id} already exists. Skipping creation.")

    # Load the Excel template
    template_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/templates/templateInvoiceFinal.xlsx')

    workbook = load_workbook(template_path)
    sheet = workbook.active
    respondent_message = f"sampel {user_data['respondent_count']} responden"
    paid_percentage_message = f"{user_data['paid_percentage']}%"
    formatted_date = datetime.strptime(user_data['date'], "%Y-%m-%d").strftime("%d %B %Y")
    date_message = f"Date: {formatted_date}"
    sheet.merge_cells('B31:E31')
    sheet.merge_cells('C16:F16')
    sheet.merge_cells('B27:E27')
    sheet.merge_cells('B28:E28')
    sheet.merge_cells('G14:H14')

    # Fill out the required cells in the Excel file with user data
    sheet['A9'] = invoice_code
    sheet['C16'] = user_data['client_name']           
    sheet['B27'] = user_data['survey_name']           
    sheet['B31'] = respondent_message      
    sheet['C18'] = user_data['address']               
    sheet['G27'] = user_data['amount']                
    sheet['C35'] = user_data['nominal_tertulis']
    sheet['C35'].font = Font(name="Times New Roman", bold=True, underline="single")
    sheet['F27'] = paid_percentage_message      
    sheet['B28'] = user_data['additional_info']       
    sheet['G14'] = date_message   
    sheet['C42'] = f": {user_data['beneficiary_bank_name']}"
    sheet['C39'] = f": {user_data['beneficiary_account_name']}"
    sheet['C40'] = f": {user_data['beneficiary_account_number']}"
    sheet['C40'].font = Font(name="Times New Roman", bold=True)
    sheet['C41'] = f": {user_data['account_currency']}"
    sheet['C43'] = f": {user_data['beneficiary_bank_address']}"
    sheet['C44'] = f": {user_data['beneficiary_swift_code']}"
    sheet['C45'] = f": {user_data['tax_id']}"               

    # Path to the image you want to add
    header_image_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/images/header.png')  
    header_img = Image(header_image_path)
    header_img.width, header_img.height = 846.6, 136  

    # Path to the image you want to add
    invoice_image_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/images/invoice.png') 
    invoice_img = Image(invoice_image_path)
    invoice_img.width, invoice_img.height = 275.9, 75.59

    # # Path to the image you want to add
    # ttd_image_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/images/ttd.png')
    # ttd_img = Image(ttd_image_path)
    # ttd_img.width, ttd_img.height = 314.83, 173.48

    # Add image to the specified cell location
    sheet.add_image(header_img, 'A1') 
    sheet.add_image(invoice_img, 'G8')

    # Generate a filename
    filename = f"{user_data['survey_name']}_invoiceFinal_{invoice_code}.xlsx"

    # Prepare the response as an Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'

    # Save workbook to the response
    workbook.save(response)
    return response

    return HttpResponse(status=405)  # Method not allowed for non-POST requests

@api_view(['POST'])
def generate_kwitansi_dp(request):
    if request.method == 'POST':
        # Parse JSON data from the request body
        data = json.loads(request.body)
    
    user_data = {
        'pembayar': data.get('pembayar', ''),
        'tujuan_pembayaran': data.get('tujuan_pembayaran', ''),
        'amount': data.get('amount', '0'),
        'nominal_tertulis' : data.get('nominal_tertulis', ''),
        'additional_info': data.get('additional_info', ''),
        'date': data.get('date', datetime.now().strftime('%Y-%m-%d')),
    }

    # Generate unique kwitansi code
    current_date = datetime.now()
    month_roman = month_to_roman(current_date.month)
    year = current_date.year
    kwitansi_number = get_next_kwitansi_number()
    kwitansi_code = f"{kwitansi_number}/IDR-KWT/{month_roman}/{year}"
    kwitansi_id = f"{kwitansi_number}/IDR-KWT/{month_roman}/{year}"
    doc = "kwitansiDP"

    formatted_date = datetime.strptime(user_data['date'], '%Y-%m-%d').strftime("Jakarta, %d %B %Y")

    # formatted_date = datetime.strptime(user_data['date'], '%Y-%m-%d').strftime("Jakarta, %d %B %Y")

    if Decimal(user_data['amount']) > Decimal('999999999999.99'):
        return HttpResponse({"error": "Amount terlalu besar untuk disimpan."}, status=400)

    # Save the data to the kwitansi_dp table
    elif not KwitansiDP.objects.filter(id=kwitansi_id).exists(): 
        KwitansiDP.objects.create(
            id=kwitansi_id,
            client_name=user_data['pembayar'],
            survey_name=user_data['tujuan_pembayaran'],
            nominal_tertulis=user_data['nominal_tertulis'],
            additional_info=user_data['additional_info'],
            amount=user_data['amount'],
            date=user_data['date'],
            doc_type=doc
        )
    else:
        print(f"Record with id {kwitansi_id} already exists. Skipping creation.")

    # load excel template for kwitansi dp
    template_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/templates/templateKwitansi.xlsx')

    workbook = load_workbook(template_path)
    sheet = workbook.active

    times_new_roman_font = Font(name="Times New Roman")
    bold_times_new_roman_font = Font(name="Times New Roman", bold=True)

    for row in sheet.iter_rows():
        for cell in row:
            cell.font = times_new_roman_font

    for cell in sheet['A']:
        cell.font = bold_times_new_roman_font

    sheet.merge_cells('A11:L11')
    sheet.merge_cells('E14:G14')
    sheet.merge_cells('E16:G16')
    # sheet.merge_cells(start_row=17, start_column=1, end_row=18, end_column=3)
    sheet.merge_cells(start_row=17, start_column=5, end_row=18, end_column=7)
    sheet.merge_cells('E19:G19')
    sheet.merge_cells('B27:E27')
    sheet.merge_cells('K27:L27')

    # fill cells with input from user_data
    sheet['A11'] = kwitansi_code
    sheet['E14'] = user_data['pembayar']
    sheet['E16'] = f"# {user_data['nominal_tertulis']} #"
    sheet['E17'] = user_data['tujuan_pembayaran']
    sheet['E19'] = user_data['additional_info']
    sheet['B27'] = user_data['amount']
    sheet['K27'] = formatted_date

    # generate file name
    filename = f"KwitansiDP_{kwitansi_code}.xlsx"
    # response as excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'

    # save  workbook to the response
    workbook.save(response)
    return response

@api_view(['POST'])
def export_existing_kwitansi_dp(request):
    if request.method == 'POST':
        # Parse JSON data from the request body
        data = json.loads(request.body)
    
    user_data = {
        'kwitansi_code': data.get('id', ''),
        'pembayar': data.get('client_name', ''),
        'tujuan_pembayaran': data.get('survey_name', ''),
        'amount': data.get('amount', '0'),
        'nominal_tertulis' : data.get('nominal_tertulis', ''),
        'additional_info': data.get('additional_info', ''),
        'date': data.get('date', datetime.now().strftime('%Y-%m-%d')),
    }


    formatted_date = datetime.strptime(user_data['date'], '%Y-%m-%d').strftime("Jakarta, %d %B %Y")


    # load excel template for kwitansi dp
    template_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/templates/templateKwitansi.xlsx')

    workbook = load_workbook(template_path)
    sheet = workbook.active

    times_new_roman_font = Font(name="Times New Roman")
    bold_times_new_roman_font = Font(name="Times New Roman", bold=True)

    for row in sheet.iter_rows():
        for cell in row:
            cell.font = times_new_roman_font

    for cell in sheet['A']:
        cell.font = bold_times_new_roman_font

    sheet.merge_cells('A11:L11')
    sheet.merge_cells('E14:G14')
    sheet.merge_cells('E16:G16')
    # sheet.merge_cells(start_row=17, start_column=1, end_row=18, end_column=3)
    sheet.merge_cells(start_row=17, start_column=5, end_row=18, end_column=7)
    sheet.merge_cells('E19:G19')
    sheet.merge_cells('B27:E27')
    sheet.merge_cells('K27:L27')

    # fill cells with input from user_data
    sheet['A11'] = user_data['kwitansi_code']
    sheet['E14'] = user_data['pembayar']
    sheet['E16'] = f"# {user_data['nominal_tertulis']} #"
    sheet['E17'] = user_data['tujuan_pembayaran']
    sheet['E19'] = user_data['additional_info']
    sheet['B27'] = user_data['amount']
    sheet['K27'] = formatted_date

    # generate file name
    filename = f"KwitansiDP_{user_data['kwitansi_code']}.xlsx"
    # response as excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'

    # save  workbook to the response
    workbook.save(response)
    return response

@api_view(['POST'])
def export_existing_kwitansi_final(request):
    if request.method == 'POST':
        # Parse JSON data from the request body
        data = json.loads(request.body)
    
    user_data = {
        'kwitansi_code': data.get('id', ''),
        'pembayar': data.get('client_name', ''),
        'tujuan_pembayaran': data.get('survey_name', ''),
        'amount': data.get('amount', '0'),
        'nominal_tertulis' : data.get('nominal_tertulis', ''),
        'additional_info': data.get('additional_info', ''),
        'date': data.get('date', datetime.now().strftime('%Y-%m-%d')),
    }


    formatted_date = datetime.strptime(user_data['date'], '%Y-%m-%d').strftime("Jakarta, %d %B %Y")


    # load excel template for kwitansi dp
    template_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/templates/templateKwitansi.xlsx')

    workbook = load_workbook(template_path)
    sheet = workbook.active

    times_new_roman_font = Font(name="Times New Roman")
    bold_times_new_roman_font = Font(name="Times New Roman", bold=True)

    for row in sheet.iter_rows():
        for cell in row:
            cell.font = times_new_roman_font

    for cell in sheet['A']:
        cell.font = bold_times_new_roman_font

    sheet.merge_cells('A11:L11')
    sheet.merge_cells('E14:G14')
    sheet.merge_cells('E16:G16')
    # sheet.merge_cells(start_row=17, start_column=1, end_row=18, end_column=3)
    sheet.merge_cells(start_row=17, start_column=5, end_row=18, end_column=7)
    sheet.merge_cells('E19:G19')
    sheet.merge_cells('B27:E27')
    sheet.merge_cells('K27:L27')

    # fill cells with input from user_data
    sheet['A11'] = user_data['kwitansi_code']
    sheet['E14'] = user_data['pembayar']
    sheet['E16'] = f"# {user_data['nominal_tertulis']} #"
    sheet['E17'] = user_data['tujuan_pembayaran']
    sheet['E19'] = user_data['additional_info']
    sheet['B27'] = user_data['amount']
    sheet['K27'] = formatted_date

    # generate file name
    filename = f"KwitansiFinal_{user_data['kwitansi_code']}.xlsx"
    # response as excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'

    # save  workbook to the response
    workbook.save(response)
    return response


@api_view(['POST'])
def generate_kwitansi_final(request):
    if request.method == 'POST':
        # Parse JSON data from the request body
        data = json.loads(request.body)
    
    user_data = {
        'pembayar': data.get('pembayar', ''),
        'tujuan_pembayaran': data.get('tujuan_pembayaran', ''),
        'amount': data.get('amount', '0'),
        'nominal_tertulis' : data.get('nominal_tertulis', ''),
        'additional_info': data.get('additional_info', ''),
        'date': data.get('date', datetime.now().strftime('%Y-%m-%d')),
    }

    # Generate unique kwitansi code
    current_date = datetime.now()
    month_roman = month_to_roman(current_date.month)
    year = current_date.year
    kwitansi_number = get_next_kwitansi_number()
    kwitansi_code = f"{kwitansi_number}/IDR-KWT/{month_roman}/{year}"
    kwitansi_id = f"{kwitansi_number}/IDR-KWT/{month_roman}/{year}"
    doc = "kwitansiFinal"

    formatted_date = datetime.strptime(user_data['date'], '%Y-%m-%d').strftime("Jakarta, %d %B %Y")

    # formatted_date = datetime.strptime(user_data['date'], '%Y-%m-%d').strftime("Jakarta, %d %B %Y")

    if Decimal(user_data['amount']) > Decimal('999999999999.99'):
        return HttpResponse({"error": "Amount terlalu besar untuk disimpan."}, status=400)

    # Save the data to the kwitansi_final table
    elif not KwitansiFinal.objects.filter(id=kwitansi_id).exists(): 
        KwitansiFinal.objects.create(
            id=kwitansi_id,
            client_name=user_data['pembayar'],
            survey_name=user_data['tujuan_pembayaran'],
            nominal_tertulis=user_data['nominal_tertulis'],
            additional_info=user_data['additional_info'],
            amount=user_data['amount'],
            date=user_data['date'],
            doc_type=doc
        )
    else:
        print(f"Record with id {kwitansi_id} already exists. Skipping creation.")

    # load excel template for kwitansi dp
    template_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/templates/templateKwitansi.xlsx')

    workbook = load_workbook(template_path)
    sheet = workbook.active

    times_new_roman_font = Font(name="Times New Roman")
    bold_times_new_roman_font = Font(name="Times New Roman", bold=True)

    for row in sheet.iter_rows():
        for cell in row:
            cell.font = times_new_roman_font

    for cell in sheet['A']:
        cell.font = bold_times_new_roman_font

    sheet.merge_cells('A11:L11')
    sheet.merge_cells('E14:G14')
    sheet.merge_cells('E16:G16')
    # sheet.merge_cells(start_row=17, start_column=1, end_row=18, end_column=3)
    sheet.merge_cells(start_row=17, start_column=5, end_row=18, end_column=7)
    sheet.merge_cells('E19:G19')
    sheet.merge_cells('B27:E27')
    sheet.merge_cells('K27:L27')

    # fill cells with input from user_data
    sheet['A11'] = kwitansi_code
    sheet['E14'] = user_data['pembayar']
    sheet['E16'] = f"# {user_data['nominal_tertulis']} #"
    sheet['E17'] = user_data['tujuan_pembayaran']
    sheet['E19'] = user_data['additional_info']
    sheet['B27'] = user_data['amount']
    sheet['K27'] = formatted_date

    # generate file name
    filename = f"KwitansiFinal_{kwitansi_code}.xlsx"
    # response as excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'

    # save  workbook to the response
    workbook.save(response)
    return response
